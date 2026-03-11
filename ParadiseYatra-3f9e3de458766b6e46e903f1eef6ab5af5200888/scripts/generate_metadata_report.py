from __future__ import annotations

import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse
from xml.etree import ElementTree as ET

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://paradiseyatra.com"
REQUEST_TIMEOUT = 20
MAX_WORKERS = 8
ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "src" / "app"
LAYOUT_FILE = APP_DIR / "layout.tsx"
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
OUTPUT_FILE = OUTPUT_DIR / "paradise-yatra-metadata-report.xlsx"
ADMIN_MANAGED_ROUTE_PATTERNS = {
    "/",
    "/blog/[slug]",
    "/fixed-departures/[slug]",
    "/package/[slug]",
    "/package/theme/[slug]",
}


@dataclass(frozen=True)
class RouteInfo:
    route_pattern: str
    source_file: str
    route_kind: str
    metadata_source: str
    static_title: str
    static_description: str
    static_keywords: str
    static_canonical: str
    static_robots: str
    notes: str


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def discover_routes(app_dir: Path) -> list[RouteInfo]:
    routes: list[RouteInfo] = []
    for page_file in sorted(app_dir.rglob("page.tsx")):
        rel_parts = page_file.relative_to(app_dir).parts[:-1]
        route_parts = [part for part in rel_parts if not (part.startswith("(") and part.endswith(")"))]
        route_pattern = "/" + "/".join(route_parts) if route_parts else "/"
        file_text = read_text(page_file)

        has_static_metadata = bool(
            re.search(r"export\s+const\s+metadata(?:\s*:\s*Metadata)?\s*=", file_text)
        )
        has_generate_metadata = bool(
            re.search(r"export\s+(?:async\s+)?function\s+generateMetadata\s*\(", file_text)
        )

        if has_static_metadata and has_generate_metadata:
            metadata_source = "static + generateMetadata"
        elif has_generate_metadata:
            metadata_source = "generateMetadata"
        elif has_static_metadata:
            metadata_source = "static metadata"
        else:
            metadata_source = "none"

        route_kind = "dynamic" if "[" in route_pattern else "static"
        routes.append(
            RouteInfo(
                route_pattern=route_pattern,
                source_file=str(page_file.relative_to(ROOT)).replace("\\", "/"),
                route_kind=route_kind,
                metadata_source=metadata_source,
                static_title=extract_literal_field(file_text, "title"),
                static_description=extract_literal_field(file_text, "description"),
                static_keywords=extract_keywords(file_text),
                static_canonical=extract_literal_field(file_text, "canonical"),
                static_robots=extract_robots(file_text),
                notes=derive_route_notes(file_text, metadata_source),
            )
        )
    return routes


def extract_literal_field(text: str, key: str) -> str:
    patterns = [
        rf"{key}\s*:\s*\"([^\"]+)\"",
        rf"{key}\s*:\s*'([^']+)'",
        rf"{key}\s*:\s*`([^`]+)`",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.DOTALL)
        if match:
            return normalize_whitespace(match.group(1))
    return ""


def extract_keywords(text: str) -> str:
    match = re.search(r"keywords\s*:\s*(\[[\s\S]*?\]|\"[^\"]+\"|'[^']+'|`[^`]+`)", text)
    if not match:
        return ""
    raw = match.group(1).strip()
    if raw.startswith("["):
        values = re.findall(r"\"([^\"]+)\"|'([^']+)'|`([^`]+)`", raw)
        flattened = [normalize_whitespace(next(item for item in group if item)) for group in values]
        return ", ".join(flattened)
    return normalize_whitespace(raw.strip("\"'`"))


def extract_robots(text: str) -> str:
    robots_match = re.search(r"robots\s*:\s*\{([\s\S]*?)\}", text)
    if not robots_match:
        return ""
    block = robots_match.group(1)
    index_value = extract_boolean_field(block, "index")
    follow_value = extract_boolean_field(block, "follow")
    values = []
    if index_value:
        values.append(f"index={index_value}")
    if follow_value:
        values.append(f"follow={follow_value}")
    return ", ".join(values)


def extract_boolean_field(text: str, key: str) -> str:
    match = re.search(rf"{key}\s*:\s*(true|false)", text)
    return match.group(1) if match else ""


def derive_route_notes(text: str, metadata_source: str) -> str:
    notes: list[str] = []
    fetch_targets = re.findall(r"fetch\(([^)]+)\)", text)
    if metadata_source == "generateMetadata" and fetch_targets:
        notes.append("Metadata is built dynamically from fetched data.")
    elif metadata_source == "none":
        notes.append("No page-level metadata export found; page inherits root metadata.")

    if "Package Not Found" in text or "Blog Post Not Found" in text:
        notes.append("Includes fallback metadata when the requested record is missing.")

    if "revalidate" in text:
        revalidate_match = re.search(r"revalidate\s*=\s*(\d+)", text)
        if revalidate_match:
            notes.append(f"Revalidates every {revalidate_match.group(1)} seconds.")

    return " ".join(notes)


def normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def load_sitemap(base_url: str) -> list[dict[str, str]]:
    response = requests.get(urljoin(base_url, "/sitemap.xml"), timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    root = ET.fromstring(response.text)
    namespace = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

    pages: list[dict[str, str]] = []
    for url_node in root.findall("sm:url", namespace):
        pages.append(
            {
                "url": url_node.findtext("sm:loc", default="", namespaces=namespace),
                "lastmod": url_node.findtext("sm:lastmod", default="", namespaces=namespace),
                "changefreq": url_node.findtext("sm:changefreq", default="", namespaces=namespace),
                "priority": url_node.findtext("sm:priority", default="", namespaces=namespace),
            }
        )
    return pages


def collect_live_urls(routes: list[RouteInfo], sitemap_entries: list[dict[str, str]], base_url: str) -> list[dict[str, str]]:
    sitemap_map = {entry["url"]: entry for entry in sitemap_entries if entry.get("url")}

    for route in routes:
        if route.route_kind != "static":
            continue
        static_url = urljoin(base_url, route.route_pattern if route.route_pattern != "/" else "")
        sitemap_map.setdefault(
            static_url,
            {
                "url": static_url,
                "lastmod": "",
                "changefreq": "",
                "priority": "",
            },
        )

    return sorted(sitemap_map.values(), key=lambda item: item["url"])


def exclude_admin_managed_routes(
    routes: list[RouteInfo], live_rows: list[dict[str, Any]]
) -> tuple[list[RouteInfo], list[dict[str, Any]]]:
    filtered_routes = [route for route in routes if route.route_pattern not in ADMIN_MANAGED_ROUTE_PATTERNS]
    filtered_live_rows = [
        row for row in live_rows if row.get("matched_route", "") not in ADMIN_MANAGED_ROUTE_PATTERNS
    ]
    return filtered_routes, filtered_live_rows


def compile_route_regex(route_pattern: str) -> re.Pattern[str]:
    if route_pattern == "/":
        return re.compile(r"^/$")

    parts = [part for part in route_pattern.strip("/").split("/") if part]
    regex_parts = []
    for part in parts:
        if part.startswith("[") and part.endswith("]"):
            regex_parts.append(r"[^/]+")
        else:
            regex_parts.append(re.escape(part))
    return re.compile(r"^/" + "/".join(regex_parts) + r"$")


def route_specificity(route_pattern: str) -> tuple[int, int]:
    parts = [part for part in route_pattern.strip("/").split("/") if part]
    static_parts = sum(1 for part in parts if not (part.startswith("[") and part.endswith("]")))
    return (len(parts), static_parts)


def match_route(path: str, route_patterns: list[tuple[RouteInfo, re.Pattern[str]]]) -> RouteInfo | None:
    matches: list[RouteInfo] = []
    for route_info, compiled in route_patterns:
        if compiled.match(path):
            matches.append(route_info)
    if not matches:
        return None
    return sorted(matches, key=lambda item: route_specificity(item.route_pattern), reverse=True)[0]


def scrape_live_page(url: str) -> dict[str, Any]:
    try:
        response = requests.get(
            url,
            timeout=REQUEST_TIMEOUT,
            headers={
                "User-Agent": "Mozilla/5.0 (Metadata Report Generator)",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        result: dict[str, Any] = {
            "url": url,
            "status_code": response.status_code,
            "final_url": response.url,
        }
        soup = BeautifulSoup(response.text, "html.parser")
        title_tag = soup.find("title")
        h1_tag = soup.find("h1")

        result.update(
            {
                "title": normalize_whitespace(title_tag.get_text()) if title_tag else "",
                "h1": normalize_whitespace(h1_tag.get_text()) if h1_tag else "",
                "description": get_meta_content(soup, "name", "description"),
                "keywords": get_meta_content(soup, "name", "keywords"),
                "canonical": get_canonical(soup),
                "robots": get_meta_content(soup, "name", "robots"),
                "googlebot": get_meta_content(soup, "name", "googlebot"),
                "og_title": get_meta_content(soup, "property", "og:title"),
                "og_description": get_meta_content(soup, "property", "og:description"),
                "og_url": get_meta_content(soup, "property", "og:url"),
                "og_type": get_meta_content(soup, "property", "og:type"),
                "og_image": get_meta_content(soup, "property", "og:image"),
                "twitter_card": get_meta_content(soup, "name", "twitter:card"),
                "twitter_title": get_meta_content(soup, "name", "twitter:title"),
                "twitter_description": get_meta_content(soup, "name", "twitter:description"),
                "twitter_image": get_meta_content(soup, "name", "twitter:image"),
            }
        )
        result["page_name"] = result["h1"] or derive_page_name_from_title(result["title"], response.url)
        result["title_length"] = len(result["title"])
        result["description_length"] = len(result["description"])
        result["issues"] = build_issue_list(result)
        return result
    except Exception as exc:  # pragma: no cover - external I/O
        return {
            "url": url,
            "status_code": "ERROR",
            "final_url": "",
            "title": "",
            "h1": "",
            "description": "",
            "keywords": "",
            "canonical": "",
            "robots": "",
            "googlebot": "",
            "og_title": "",
            "og_description": "",
            "og_url": "",
            "og_type": "",
            "og_image": "",
            "twitter_card": "",
            "twitter_title": "",
            "twitter_description": "",
            "twitter_image": "",
            "page_name": derive_page_name_from_title("", url),
            "title_length": 0,
            "description_length": 0,
            "issues": f"Fetch error: {exc}",
        }


def get_meta_content(soup: BeautifulSoup, attr_name: str, attr_value: str) -> str:
    tag = soup.find("meta", attrs={attr_name: attr_value})
    if not tag:
        return ""
    content = tag.get("content", "")
    return normalize_whitespace(content)


def get_canonical(soup: BeautifulSoup) -> str:
    tag = soup.find("link", rel=lambda rel: rel and "canonical" in rel)
    if not tag:
        return ""
    href = tag.get("href", "")
    return normalize_whitespace(href)


def derive_page_name_from_title(title: str, url: str) -> str:
    if title:
        return title.split("|")[0].strip()

    parsed = urlparse(url)
    path = parsed.path.strip("/")
    if not path:
        return "Home"
    return " ".join(part.capitalize() for part in path.split("/")[-1].split("-"))


def build_issue_list(page: dict[str, Any]) -> str:
    issues: list[str] = []
    if not page.get("title"):
        issues.append("Missing title")
    if not page.get("description"):
        issues.append("Missing description")
    if not page.get("canonical"):
        issues.append("Missing canonical")
    if not page.get("og_title"):
        issues.append("Missing og:title")
    if not page.get("og_description"):
        issues.append("Missing og:description")
    if not page.get("twitter_title"):
        issues.append("Missing twitter:title")
    if not page.get("twitter_description"):
        issues.append("Missing twitter:description")
    robots = str(page.get("robots", "")).lower()
    if "noindex" in robots:
        issues.append("Noindex")
    return "; ".join(issues)


def scrape_live_pages(urls: list[str]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape_live_page, url): url for url in urls}
        for future in as_completed(futures):
            results.append(future.result())
    return sorted(results, key=lambda item: item["url"])


def extract_global_metadata(layout_text: str) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = [
        ("source_file", "src/app/layout.tsx"),
        ("metadataBase", extract_new_url_argument(layout_text)),
        ("icons.icon", extract_literal_field(layout_text, "icon")),
        ("icons.shortcut", extract_literal_field(layout_text, "shortcut")),
        ("icons.apple", extract_literal_field(layout_text, "apple")),
        ("verification.google", extract_global_key_value(layout_text, "google")),
        ("other.theme-color", extract_global_key_value(layout_text, "theme-color")),
        ("other.sitemap", extract_global_key_value(layout_text, "sitemap")),
        ("other.Content-Security-Policy", extract_global_key_value(layout_text, "Content-Security-Policy")),
    ]
    return rows


def extract_new_url_argument(text: str) -> str:
    match = re.search(r"metadataBase\s*:\s*new URL\(([\s\S]*?)\)", text)
    return normalize_whitespace(match.group(1)) if match else ""


def extract_global_key_value(text: str, key: str) -> str:
    patterns = [
        rf"\"{re.escape(key)}\"\s*:\s*\"([^\"]+)\"",
        rf"'{re.escape(key)}'\s*:\s*'([^']+)'",
        rf"{re.escape(key)}\s*:\s*\"([^\"]+)\"",
        rf"{re.escape(key)}\s*:\s*'([^']+)'",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return normalize_whitespace(match.group(1))
    return ""


def write_workbook(
    live_rows: list[dict[str, Any]],
    routes: list[RouteInfo],
    global_metadata_rows: list[tuple[str, str]],
    generated_at: str,
) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    live_ws = workbook.create_sheet("Live Metadata")
    route_ws = workbook.create_sheet("Route Audit")
    global_ws = workbook.create_sheet("Global Metadata")

    write_summary_sheet(summary_ws, generated_at)
    write_live_metadata_sheet(live_ws, live_rows)
    write_route_audit_sheet(route_ws, routes, live_rows)
    write_global_metadata_sheet(global_ws, global_metadata_rows)

    for worksheet in workbook.worksheets:
        style_worksheet(worksheet)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)


def write_summary_sheet(sheet, generated_at: str) -> None:
    rows = [
        ("Report", "Paradise Yatra Metadata Report (Admin-managed pages excluded)"),
        ("Generated At (UTC)", generated_at),
        ("Base URL", BASE_URL),
        ("Total App Routes", "=COUNTA('Route Audit'!A:A)-1"),
        ("Total Live URLs Audited", "=COUNTA('Live Metadata'!A:A)-1"),
        ("Routes Using generateMetadata", '=COUNTIF(\'Route Audit\'!D:D,"*generateMetadata*")'),
        ("Routes With No Page Metadata", '=COUNTIF(\'Route Audit\'!D:D,"none")'),
        ("Live URLs Missing Description", '=COUNTIF(\'Live Metadata\'!AC:AC,"*Missing description*")'),
        ("Live URLs Missing Canonical", '=COUNTIF(\'Live Metadata\'!AC:AC,"*Missing canonical*")'),
        ("Live URLs Marked Noindex", '=COUNTIF(\'Live Metadata\'!AC:AC,"*Noindex*")'),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 80


def write_live_metadata_sheet(sheet, live_rows: list[dict[str, Any]]) -> None:
    headers = [
        "Page Name",
        "URL",
        "Matched Route",
        "Route File",
        "Source Type",
        "Sitemap Last Modified",
        "Sitemap Changefreq",
        "Sitemap Priority",
        "HTTP Status",
        "Final URL",
        "Title",
        "Title Length",
        "Meta Description",
        "Description Length",
        "Meta Keywords",
        "Canonical",
        "Robots",
        "Googlebot",
        "OG Title",
        "OG Description",
        "OG URL",
        "OG Type",
        "OG Image",
        "Twitter Card",
        "Twitter Title",
        "Twitter Description",
        "Twitter Image",
        "H1",
        "Issues",
    ]
    sheet.append(headers)
    for row in live_rows:
        sheet.append(
            [
                row.get("page_name", ""),
                row.get("url", ""),
                row.get("matched_route", ""),
                row.get("route_file", ""),
                row.get("source_type", ""),
                row.get("sitemap_lastmod", ""),
                row.get("sitemap_changefreq", ""),
                row.get("sitemap_priority", ""),
                row.get("status_code", ""),
                row.get("final_url", ""),
                row.get("title", ""),
                row.get("title_length", 0),
                row.get("description", ""),
                row.get("description_length", 0),
                row.get("keywords", ""),
                row.get("canonical", ""),
                row.get("robots", ""),
                row.get("googlebot", ""),
                row.get("og_title", ""),
                row.get("og_description", ""),
                row.get("og_url", ""),
                row.get("og_type", ""),
                row.get("og_image", ""),
                row.get("twitter_card", ""),
                row.get("twitter_title", ""),
                row.get("twitter_description", ""),
                row.get("twitter_image", ""),
                row.get("h1", ""),
                row.get("issues", ""),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_route_audit_sheet(sheet, routes: list[RouteInfo], live_rows: list[dict[str, Any]]) -> None:
    headers = [
        "Route Pattern",
        "Route Kind",
        "Source File",
        "Metadata Source",
        "Sitemap URL Count",
        "Sample Live URL",
        "Static Title",
        "Static Description",
        "Static Keywords",
        "Static Canonical",
        "Static Robots",
        "Notes",
    ]
    sheet.append(headers)

    route_to_urls: dict[str, list[dict[str, Any]]] = {}
    for row in live_rows:
        matched_route = row.get("matched_route", "")
        if matched_route:
            route_to_urls.setdefault(matched_route, []).append(row)

    for route in routes:
        matched_rows = route_to_urls.get(route.route_pattern, [])
        sheet.append(
            [
                route.route_pattern,
                route.route_kind,
                route.source_file,
                route.metadata_source,
                len(matched_rows),
                matched_rows[0]["url"] if matched_rows else "",
                route.static_title,
                route.static_description,
                route.static_keywords,
                route.static_canonical,
                route.static_robots,
                route.notes,
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_global_metadata_sheet(sheet, rows: list[tuple[str, str]]) -> None:
    sheet.append(["Key", "Value"])
    for key, value in rows:
        sheet.append([key, value])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def style_worksheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)


def enrich_live_rows(
    scraped_rows: list[dict[str, Any]],
    routes: list[RouteInfo],
    sitemap_entries: list[dict[str, str]],
) -> list[dict[str, Any]]:
    route_patterns = [(route, compile_route_regex(route.route_pattern)) for route in routes]
    sitemap_lookup = {entry["url"]: entry for entry in sitemap_entries}

    enriched: list[dict[str, Any]] = []
    for row in scraped_rows:
        path = urlparse(row["url"]).path or "/"
        matched_route = match_route(path, route_patterns)
        sitemap_entry = sitemap_lookup.get(row["url"], {})
        source_type = "sitemap + app route" if sitemap_entry and matched_route else "sitemap" if sitemap_entry else "app route"
        enriched.append(
            {
                **row,
                "matched_route": matched_route.route_pattern if matched_route else "",
                "route_file": matched_route.source_file if matched_route else "",
                "source_type": source_type,
                "sitemap_lastmod": sitemap_entry.get("lastmod", ""),
                "sitemap_changefreq": sitemap_entry.get("changefreq", ""),
                "sitemap_priority": sitemap_entry.get("priority", ""),
            }
        )
    return enriched


def main() -> None:
    routes = discover_routes(APP_DIR)
    sitemap_entries = load_sitemap(BASE_URL)
    live_url_entries = collect_live_urls(routes, sitemap_entries, BASE_URL)
    live_urls = [entry["url"] for entry in live_url_entries]
    scraped_rows = scrape_live_pages(live_urls)
    enriched_rows = enrich_live_rows(scraped_rows, routes, live_url_entries)
    routes, enriched_rows = exclude_admin_managed_routes(routes, enriched_rows)
    layout_rows = extract_global_metadata(read_text(LAYOUT_FILE))
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    write_workbook(enriched_rows, routes, layout_rows, generated_at)
    print(f"Created {OUTPUT_FILE}")
    print(f"Routes audited: {len(routes)}")
    print(f"Live URLs audited: {len(enriched_rows)}")


if __name__ == "__main__":
    main()
